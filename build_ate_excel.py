"""
build_ate_excel.py
------------------
Reads ATE extraction CSV and produces an Excel workbook with native charts.
Excel-on-Windows compatible — series titles use StrRef cell references,
scatter charts use proper xVal/yVal numRef, colours use 6-char srgbClr.

Tabs: RAW DATA | SUMMARY STATS | BOLTZMANN | RUN TREND | BOX PLOT | HISTOGRAM

Usage:
    python build_ate_excel.py --input results.csv --output ATE_Report.xlsx

Requirements:
    pip install openpyxl pandas
"""

import argparse, os, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, BarChart, LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef, StrData, StrVal

# ── Palette & constants ───────────────────────────────────────────────────────
PALETTE      = {"COLD": "3A7ABF", "ROOM": "E040A0", "HOT": "00A898"}
HEADER_FILL  = "1F3864"
ALT_ROW_FILL = "F2F7FB"
WHITE        = "FFFFFF"
BORDER_COLOR = "BFBFBF"
TEMP_ORDER   = ["COLD", "ROOM", "HOT"]
TEMP_AXIS_MAP = {"COLD": -40, "ROOM": 25, "HOT": 85}
VALUE_COL    = "extracted_value"

# ── Style helpers ─────────────────────────────────────────────────────────────

def tborder():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, color=WHITE):
    return Font(name="Arial", size=size, bold=True, color=color)

def bfont(size=9, bold=False):
    return Font(name="Arial", size=size, bold=bold)

def sfill(h):
    return PatternFill("solid", fgColor=h)

def write_header(ws, row, cols):
    for c, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = hfont(); cell.fill = sfill(HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = tborder()

def write_row(ws, row_i, vals, alt=False):
    bg = ALT_ROW_FILL if alt else WHITE
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row_i, column=c, value=val)
        cell.font = bfont(); cell.fill = sfill(bg)
        cell.border = tborder()
        cell.alignment = Alignment(horizontal="center", vertical="center")

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Domain helpers ────────────────────────────────────────────────────────────

def detect_temp(folder):
    u = folder.upper()
    for k in ("COLD", "HOT", "ROOM"):
        if k in u: return k
    return "ROOM"

def detect_param(folder):
    return re.sub(r"_(COLD|HOT|ROOM)$", "", folder, flags=re.IGNORECASE)

def safe_run(val, fallback):
    try:
        if pd.notna(val): return int(val)
    except (TypeError, ValueError): pass
    return fallback

def col_addr(ws_title, col, row):
    """Return an absolute Excel cell address string like 'BOLTZMANN'!$B$5"""
    safe = ws_title.replace("'", "''")
    return f"'{safe}'!${get_column_letter(col)}${row}"

def cell_series_label(ws_title, col, row, display_value):
    """
    Build a SeriesLabel that points to a cell — Excel-safe.
    display_value is cached so the chart renders without recalculation.
    """
    f     = col_addr(ws_title, col, row)
    cache = StrData(pt=[StrVal(idx=0, v=str(display_value))])
    return SeriesLabel(strRef=StrRef(f=f, strCache=cache))

def apply_line_colour(ser, hex6):
    """Set line colour on a series (works for line & scatter)."""
    ser.graphicalProperties.line.solidFill = hex6

def apply_fill_colour(ser, hex6):
    """Set fill colour on a series (works for bar)."""
    ser.graphicalProperties.solidFill = hex6

def apply_marker_colour(ser, hex6):
    """Set marker fill + outline colour."""
    ser.marker.graphicalProperties.solidFill      = hex6
    ser.marker.graphicalProperties.line.solidFill = hex6


# ── RAW DATA ──────────────────────────────────────────────────────────────────

def build_raw_data(wb, df):
    ws = wb.create_sheet("RAW DATA")
    ws.freeze_panes = "A2"
    cols    = ["test_folder","temp_folder","run_number",VALUE_COL,"temp_cond","parameter"]
    headers = ["Test Folder","Temp Folder","Run #","Extracted Value","Temp Condition","Parameter"]
    write_header(ws, 1, headers)
    for i, (_, row) in enumerate(df[cols].iterrows()):
        write_row(ws, i+2, list(row), alt=(i%2==1))
        cell = ws.cell(row=i+2, column=5)
        cell.fill = sfill(PALETTE.get(row["temp_cond"], "AAAAAA"))
        cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    set_widths(ws, [14,20,8,18,14,16])


# ── SUMMARY STATS ─────────────────────────────────────────────────────────────

def build_summary(wb, df):
    ws = wb.create_sheet("SUMMARY STATS")
    ws.freeze_panes = "A2"
    params     = sorted(df["parameter"].unique())
    conditions = [c for c in TEMP_ORDER if c in df["temp_cond"].unique()]
    write_header(ws, 1, ["Parameter","Condition","Mean","Std Dev","Min","Max","Count"])
    r = 2
    for param in params:
        for cond in conditions:
            sub = df[(df["parameter"]==param)&(df["temp_cond"]==cond)][VALUE_COL].dropna()
            if sub.empty: continue
            vals = [param, cond, sub.mean(), sub.std(), sub.min(), sub.max(), len(sub)]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = bfont(); cell.border = tborder()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = sfill(ALT_ROW_FILL if r%2==0 else WHITE)
                if c == 2:
                    cell.fill = sfill(PALETTE.get(cond,"AAAAAA"))
                    cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
                if c in (3,4,5,6):
                    cell.number_format = "0.000000E+00"
            r += 1
    set_widths(ws, [18,12,16,16,16,16,8])


# ── BOLTZMANN SCATTER ─────────────────────────────────────────────────────────

def build_boltzmann(wb, df, temp_map=None):
    """
    Native scatter mirroring the matplotlib Boltzmann plot:
    x-axis is Temperature (°C), each condition is a cluster at its temperature
    position with a small run-number-based jitter so overlapping points spread.
    temp_map overrides TEMP_AXIS_MAP for the x positions.
    """
    tmap = temp_map or TEMP_AXIS_MAP
    ws = wb.create_sheet("BOLTZMANN")
    ws["A1"] = "Boltzmann Scatter — Value by Temperature Condition"
    ws["A1"].font = hfont(size=13, color=HEADER_FILL)

    params       = sorted(df["parameter"].unique())
    conditions   = [c for c in TEMP_ORDER if c in df["temp_cond"].unique()]
    jitter_span  = 1.2
    used_temps   = [tmap.get(c, 25) for c in conditions]
    x_axis_min   = (min(used_temps) - 15) if used_temps else -55
    x_axis_max   = (max(used_temps) + 15) if used_temps else 100
    data_col     = 1
    chart_row    = 3

    for param in params:
        lbl_row = chart_row
        hdr_row = chart_row + 1
        ws.cell(row=lbl_row, column=data_col, value=param).font = hfont(size=9, color=HEADER_FILL)

        col_off      = data_col
        series_meta  = []   # (cond, x_col, y_col, title_row, first_data_row, last_data_row)

        for cond in conditions:
            sub = df[(df["parameter"]==param)&(df["temp_cond"]==cond)].sort_values("run_number")
            if sub.empty: continue

            x_col, y_col = col_off, col_off+1
            title_row    = hdr_row

            # Write the condition name into the title cell (chart will ref this)
            ws.cell(row=title_row, column=y_col, value=cond)
            ws.cell(row=title_row, column=x_col, value=f"{cond} \u00b0C")

            # Compute a small jitter around the condition's base temperature so
            # overlapping points spread visually (mirrors plot_boltzmann).
            x_base    = tmap.get(cond, 25)
            runs      = pd.to_numeric(sub["run_number"], errors="coerce")
            run_mean  = runs.mean()
            run_std   = runs.std() or 1
            scale     = max(run_std, 1)

            first_data = hdr_row + 1
            dr = first_data
            for seq, (_, r) in enumerate(sub.iterrows(), 1):
                run_val = safe_run(r["run_number"], seq)
                offset  = ((run_val - run_mean) / scale) * jitter_span if pd.notna(run_mean) else 0
                ws.cell(row=dr, column=x_col, value=float(x_base + offset))
                ws.cell(row=dr, column=y_col, value=float(r[VALUE_COL]))
                dr += 1

            series_meta.append((cond, x_col, y_col, title_row, first_data, dr-1))
            col_off += 3

        # Build chart
        chart = ScatterChart()
        chart.title           = f"Boltzmann \u2014 {param}"
        chart.style           = 2
        chart.x_axis.title    = "Temperature (\u00b0C)"
        chart.y_axis.title    = f"{param}  (Extracted Value)"
        chart.x_axis.numFmt   = "0"
        chart.x_axis.scaling.min = x_axis_min
        chart.x_axis.scaling.max = x_axis_max
        chart.legend.position = "r"
        chart.width = 22; chart.height = 13

        for cond, x_col, y_col, title_row, first_row, last_row in series_meta:
            xvals = Reference(ws, min_col=x_col, min_row=first_row, max_row=last_row)
            yvals = Reference(ws, min_col=y_col, min_row=first_row, max_row=last_row)
            ser   = Series(yvals, xvals)
            # Cell-reference title — the only form Excel reliably accepts
            ser.title = cell_series_label(ws.title, y_col, title_row, cond)

            ser.marker.symbol = "circle"
            ser.marker.size   = 5
            # Line between points: none (scatter, not line-scatter)
            ser.graphicalProperties.line.noFill = True
            hex6 = PALETTE.get(cond, "AAAAAA")
            apply_marker_colour(ser, hex6)
            chart.series.append(ser)

        ws.add_chart(chart, f"{get_column_letter(col_off+1)}{chart_row}")
        chart_row += 22

    set_widths(ws, [10]*50)


# ── RUN TREND ─────────────────────────────────────────────────────────────────

def build_trend(wb, df):
    ws = wb.create_sheet("RUN TREND")
    ws["A1"] = "Run Trend \u2014 Value vs. Run Number"
    ws["A1"].font = hfont(size=13, color=HEADER_FILL)

    params     = sorted(df["parameter"].unique())
    conditions = [c for c in TEMP_ORDER if c in df["temp_cond"].unique()]
    data_col   = 1
    chart_row  = 3

    for param in params:
        lbl_row = chart_row
        hdr_row = chart_row + 1
        ws.cell(row=lbl_row, column=data_col, value=param).font = hfont(size=9, color=HEADER_FILL)

        col_off     = data_col
        series_meta = []

        for cond in conditions:
            sub = df[(df["parameter"]==param)&(df["temp_cond"]==cond)].sort_values("run_number")
            if sub.empty: continue

            y_col      = col_off
            title_row  = hdr_row
            ws.cell(row=title_row, column=y_col, value=cond)

            first_data = hdr_row + 1
            dr = first_data
            for seq, (_, r) in enumerate(sub.iterrows(), 1):
                ws.cell(row=dr, column=y_col, value=float(r[VALUE_COL]))
                dr += 1

            series_meta.append((cond, y_col, title_row, first_data, dr-1))
            col_off += 2

        chart = LineChart()
        chart.title           = f"Run Trend \u2014 {param}"
        chart.style           = 2
        chart.x_axis.title    = "Run #"
        chart.y_axis.title    = f"{param}  (Extracted Value)"
        chart.legend.position = "r"
        chart.width = 22; chart.height = 13

        for cond, y_col, title_row, first_row, last_row in series_meta:
            yvals = Reference(ws, min_col=y_col, min_row=first_row, max_row=last_row)
            ser   = Series(yvals)
            ser.title = cell_series_label(ws.title, y_col, title_row, cond)
            hex6 = PALETTE.get(cond, "AAAAAA")
            apply_line_colour(ser, hex6)
            ser.graphicalProperties.line.width = 20000   # ~2.2 pt
            ser.marker.symbol = "circle"; ser.marker.size = 4
            apply_marker_colour(ser, hex6)
            chart.series.append(ser)

        ws.add_chart(chart, f"{get_column_letter(col_off+1)}{chart_row}")
        chart_row += 22

    set_widths(ws, [10]*40)


# ── BOX PLOT (mean bar) ───────────────────────────────────────────────────────

def build_boxplot(wb, df):
    ws = wb.create_sheet("BOX PLOT")
    ws["A1"] = "Distribution Summary \u2014 Mean per Condition"
    ws["A1"].font = hfont(size=13, color=HEADER_FILL)

    params     = sorted(df["parameter"].unique())
    conditions = [c for c in TEMP_ORDER if c in df["temp_cond"].unique()]

    tbl_start = 3
    write_header(ws, tbl_start, ["Parameter","Condition","Mean","Std Dev","Min","Max","Count"])

    dr = tbl_start + 1
    for param in params:
        for cond in conditions:
            sub = df[(df["parameter"]==param)&(df["temp_cond"]==cond)][VALUE_COL].dropna()
            if sub.empty: continue
            vals = [param, cond, sub.mean(), sub.std(), sub.min(), sub.max(), len(sub)]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row=dr, column=c, value=val)
                cell.font = bfont(); cell.border = tborder()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = sfill(ALT_ROW_FILL if dr%2==0 else WHITE)
                if c == 2:
                    cell.fill = sfill(PALETTE.get(cond,"AAAAAA"))
                    cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
                if c in (3,4,5,6):
                    cell.number_format = "0.000000E+00"
            dr += 1

    tbl_end = dr - 1
    chart_col = 9
    chart_row = tbl_start

    for param in params:
        param_rows = [r for r in range(tbl_start+1, tbl_end+1)
                      if ws.cell(row=r, column=1).value == param]
        if not param_rows: continue

        chart = BarChart()
        chart.type = "col"; chart.style = 2
        chart.title        = f"Mean \u2014 {param}"
        chart.y_axis.title = f"{param}  (Mean Value)"
        chart.x_axis.title = "Temp Condition"
        chart.legend.position = "r"
        chart.width = 16; chart.height = 13

        cats = Reference(ws, min_col=2, min_row=param_rows[0], max_row=param_rows[-1])
        data = Reference(ws, min_col=3, min_row=tbl_start,     max_row=param_rows[-1])
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, f"{get_column_letter(chart_col)}{chart_row}")
        chart_row += 22

    set_widths(ws, [16,10,16,16,16,16,8])


# ── HISTOGRAM ─────────────────────────────────────────────────────────────────

def build_histogram(wb, df):
    ws = wb.create_sheet("HISTOGRAM")
    ws["A1"] = "Histogram \u2014 Value Distribution by Temp Condition"
    ws["A1"].font = hfont(size=13, color=HEADER_FILL)

    params     = sorted(df["parameter"].unique())
    conditions = [c for c in TEMP_ORDER if c in df["temp_cond"].unique()]
    N_BINS     = 15
    data_col   = 1
    chart_row  = 3

    for param in params:
        sub_all = df[df["parameter"]==param][VALUE_COL].dropna()
        if sub_all.empty: continue

        _, bin_edges = pd.cut(sub_all, bins=N_BINS, retbins=True)
        bin_labels   = [f"{bin_edges[i]:.3g}\u2013{bin_edges[i+1]:.3g}" for i in range(N_BINS)]

        lbl_row = chart_row
        hdr_row = chart_row + 1
        ws.cell(row=lbl_row, column=data_col, value=param).font = hfont(size=9, color=HEADER_FILL)
        ws.cell(row=hdr_row, column=data_col, value="Bin Range")
        for b, lbl in enumerate(bin_labels):
            ws.cell(row=hdr_row+1+b, column=data_col, value=lbl)

        col_off = data_col + 1
        chart = BarChart()
        chart.type = "col"; chart.style = 2
        chart.title        = f"Histogram \u2014 {param}"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Value Range"
        chart.legend.position = "r"
        chart.grouping = "clustered"
        chart.width = 24; chart.height = 13

        cats = Reference(ws, min_col=data_col, min_row=hdr_row+1, max_row=hdr_row+N_BINS)

        for cond in conditions:
            sub = df[(df["parameter"]==param)&(df["temp_cond"]==cond)][VALUE_COL].dropna()
            if sub.empty: continue
            binned = pd.cut(sub, bins=bin_edges, labels=False, include_lowest=True)
            counts = [int((binned==i).sum()) for i in range(N_BINS)]

            title_row = hdr_row
            ws.cell(row=title_row, column=col_off, value=cond)
            for b, cnt in enumerate(counts):
                ws.cell(row=hdr_row+1+b, column=col_off, value=cnt)

            data_ref = Reference(ws, min_col=col_off, min_row=hdr_row, max_row=hdr_row+N_BINS)
            ser = Series(data_ref, title_from_data=True)
            hex6 = PALETTE.get(cond, "AAAAAA")
            apply_fill_colour(ser, hex6)
            apply_line_colour(ser, hex6)
            chart.series.append(ser)
            col_off += 2

        chart.set_categories(cats)
        ws.add_chart(chart, f"{get_column_letter(col_off+1)}{chart_row}")
        chart_row += N_BINS + 7

    set_widths(ws, [18]+[10]*30)


# ── Main ──────────────────────────────────────────────────────────────────────

def enrich(df):
    """
    Coerce numerics and add the temp_cond / parameter columns the sheet
    builders rely on. Idempotent; operates on a copy.
    """
    required = {"test_folder", "temp_folder", "run_number", VALUE_COL}
    missing  = required - set(df.columns)
    if missing:
        raise ValueError(f"Input is missing columns: {missing}\n"
                         f"       Expected columns: {required}")

    df = df.copy()
    df[VALUE_COL]    = pd.to_numeric(df[VALUE_COL],    errors="coerce")
    df["run_number"] = pd.to_numeric(df["run_number"], errors="coerce")
    df["temp_cond"]  = df["temp_folder"].apply(detect_temp)
    df["parameter"]  = df["temp_folder"].apply(detect_param)
    df.dropna(subset=[VALUE_COL], inplace=True)
    return df


def load_and_enrich(csv_path):
    """Read the extraction CSV and enrich it for report building."""
    return enrich(pd.read_csv(csv_path))


def build_workbook(df, output, log=print, progress_callback=None, temp_map=None):
    """
    Build the full ATE workbook from an (enriched or raw) DataFrame and save
    it to ``output``. Returns the absolute output path.

    progress_callback(done, total) fires after each sheet.
    temp_map overrides the Boltzmann sheet's temperature x positions.
    """
    df = enrich(df)

    log(f"  {len(df)} rows | {df['parameter'].nunique()} params | "
        f"{df['temp_cond'].nunique()} conditions | "
        f"{df['test_folder'].nunique()} test iterations")

    sheets = [
        ("RAW DATA",      build_raw_data),
        ("SUMMARY STATS", build_summary),
        ("BOLTZMANN",     build_boltzmann),
        ("RUN TREND",     build_trend),
        ("BOX PLOT",      build_boxplot),
        ("HISTOGRAM",     build_histogram),
    ]

    wb = Workbook()
    wb.remove(wb.active)
    log("Building sheets...")
    total = len(sheets)
    for i, (name, builder) in enumerate(sheets, 1):
        log(f"  [{i}/{total}] {name}")
        if builder is build_boltzmann:
            builder(wb, df, temp_map=temp_map)
        else:
            builder(wb, df)
        if progress_callback is not None:
            progress_callback(i, total)

    out = os.path.expanduser(output)
    out_dir = os.path.dirname(os.path.abspath(out))
    os.makedirs(out_dir, exist_ok=True)
    wb.save(out)
    log(f"\nSaved \u2192 {out}")
    return os.path.abspath(out)


def main():
    parser = argparse.ArgumentParser(description="Build ATE Excel report with native charts.")
    parser.add_argument("--input",  default="results.csv",     help="Extraction CSV")
    parser.add_argument("--output", default="ATE_Report.xlsx", help="Output .xlsx path")
    args = parser.parse_args()

    print(f"Loading: {args.input}")
    df = load_and_enrich(args.input)
    build_workbook(df, args.output)

if __name__ == "__main__":
    main()
